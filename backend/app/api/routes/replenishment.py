"""补货建议模块 — 从 insights.py 拆出"""
from fastapi import APIRouter
from app.core.database import get_db
from app.core.response import ok
from app.core.sales_utils import calc_sales, rolling_predict
from datetime import datetime, timedelta, timezone
import json, os, logging

router = APIRouter(prefix="/api/insights", tags=["insights"])

logger = logging.getLogger("replenishment")


@router.get('/replenishment')
def get_replenishment_suggestions(days: int = 28, source: str = '', mode: str = 'bbcc', channel: str = 'jd', page: int = 0, page_size: int = 0, search: str = '', db = get_db()):
    """补货建议，支持 days=7/14/28 切换，mode=bbcc/traditional 切换模型"""
    # 尝试读取缓存
    from app.core.replenishment_cache import get_cached, set_cache
    cached, hit = get_cached(mode, channel, days, db)
    if hit:
        # 缓存存储格式为 {"data": suggestions}，解包后统一返回 ok 格式
        _sug = cached.get("data", []) if isinstance(cached, dict) else cached
        if search:
            _sq = search.lower()
            _sug = [s for s in _sug if _sq in str(s.get('sku','')).lower() or _sq in str(s.get('product_name','')).lower() or _sq in str(s.get('barcode','')).lower()]
        if page > 0 and page_size > 0:
            _total = len(_sug)
            return ok({"items": _sug[(page - 1) * page_size: page * page_size], "total": _total, "page": page, "page_size": page_size})
        return ok(_sug)

    try: db.table("alerts").update({"status": "inactive"}).eq("alert_type", "storage_fee").eq("status", "active").execute()
    except Exception as e: logger.warning(f"clear storage_fee alerts: {e}")

    # B仓超储预警
    try:
        pos = db.table("purchase_orders").select("*").execute().data or []
        now = datetime.now(timezone.utc)
        for po in pos:
            ad = po.get("arrival_date", "")
            if not ad or po.get("status") == "completed": continue
            try: days_stored = (now - datetime.strptime(ad[:10], "%Y-%m-%d")).days
            except Exception as e: logger.warning(f"parse date {ad}: {e}"); continue
            sku = po.get("sku", "")
            existing = db.table("alerts").select("id").eq("alert_type","b_storage_warn").eq("related_sku",sku).eq("status","active").execute().data
            if days_stored >= 11 and days_stored < 15:
                if not existing: db.table("alerts").insert({"alert_type":"b_storage_warn","title":f"B仓即将超免费期: {po.get('product_name',sku)}",
                    "description":f"入库已{days_stored}天，即将超B仓15天免费期","severity":"info","source":"replenishment_engine","related_sku":sku,"status":"active","channel":channel}).execute()
            elif days_stored >= 15 and days_stored < 20:
                if not existing: db.table("alerts").insert({"alert_type":"b_storage_warn","title":f"B仓超免费期: {po.get('product_name',sku)}",
                    "description":f"入库已{days_stored}天，超B仓15天免费期，产生仓储费","severity":"warning","source":"replenishment_engine","related_sku":sku,"status":"active","channel":channel}).execute()
            elif days_stored >= 20:
                if existing: db.table("alerts").update({"severity":"error","description":f"入库已{days_stored}天，远超B仓15天免费期，仓储费持续累计"}).eq("id",existing[0]["id"]).execute()
                else: db.table("alerts").insert({"alert_type":"b_storage_warn","title":f"B仓严重超期: {po.get('product_name',sku)}",
                    "description":f"入库已{days_stored}天，远超B仓15天免费期，仓储费持续累计","severity":"error","source":"replenishment_engine","related_sku":sku,"status":"active","channel":channel}).execute()
            elif days_stored >= 11 and existing:
                db.table("alerts").update({"description":f"入库已{days_stored}天，即将超B仓15天免费期"}).eq("id",existing[0]["id"]).execute()
    except Exception as e: logger.warning(f"B仓超储预警: {e}")

    cfg_rows = db.table("replenishment_config").select("*").eq("channel", channel).execute().data
    raw = {r['key']: r['value'] for r in cfg_rows}
    cfg = {}
    prefix = f'mode_{mode}_'
    for k, v in raw.items():
        if k.startswith(prefix): cfg[k[len(prefix):]] = v
    for k, v in raw.items():
        if not k.startswith('mode_') and k not in cfg: cfg[k] = v

    products = {p["sku"]: p for p in db.table("products").select("*").eq("deleted_at", "").execute().data}

    # 构建 barcode 映射，用于日销复合 key（sku|barcode），提高匹配精度
    sku_barcode_map = {sku: p.get('barcode', '') or '' for sku, p in products.items()}

    # 统一数据源：快照(历史) + 当天orders(实时)，消除重复计算
    from app.core.sales_utils import load_daily_sales, calc_sales_from_daily
    # 只加载当天订单（快照已含历史），从 2 万行 → 几十行
    today = datetime.now(timezone.utc).strftime('%Y-%m-%d')
    orders = db.table("orders").select("*").eq("channel", channel).gte("ordered_at", today).execute().data
    orders = [o for o in orders if not (o.get("deleted_at") or "")]

    # 三周期日销：一次遍历算 3 个窗口
    # 全量日销(所有SKU)——补货必须算所有SKU(含缺货/无库存的),
    # 不能只加载库存SKU, 否则卖光缺货的SKU会被遗漏(最需要补货)
    if mode == 'bbcc':
        # BBCC 口径: 全国 C 仓(warehouse_type='platform')销量合计——B仓是调拨仓不产生零售,
        # 自有仓集货也不计入; 用逐仓快照筛选 C 仓名 (与传统模式 C 仓逐仓口径同源)
        from app.core.sales_utils import load_daily_sales_grouped, calc_sales_from_daily as _csfd
        _c140, _c1wh = load_daily_sales_grouped(28, db, sku_barcode_map=sku_barcode_map, channel=channel)
        _c_whs = {r.get('warehouse') for r in db.table('inventory').select('warehouse').eq('warehouse_type', 'platform').eq('channel', channel).execute().data if r.get('warehouse')}
        daily_28 = {}
        for _wk, _wd in _c1wh.items():
            # _c1wh key = key|warehouse (key 可能含 sku|barcode), 拆出仓名
            _base = _wk.rsplit('|', 1)[0]
            _w = _wk.rsplit('|', 1)[1]
            if _w in _c_whs:  # 只算明确C仓名——源头已杜绝无仓订单, 未知归属不入BBCC需求
                _agg = daily_28.setdefault(_base, {})
                for _dd, _qq in _wd.items():
                    _agg[_dd] = _agg.get(_dd, 0) + _qq
    else:
        daily_28 = load_daily_sales(60, db, sku_barcode_map=sku_barcode_map, channel=channel)
    # 7/14/28 窗口日均不变(取前N天); 加 sales_60 作 28 天趋势图标的对比基准(曾缺近28趋势)
    sales_7 = calc_sales_from_daily(daily_28, 7, orders=orders, sku_barcode_map=sku_barcode_map)
    sales_14 = calc_sales_from_daily(daily_28, 14, orders=orders, sku_barcode_map=sku_barcode_map)
    sales_28 = calc_sales_from_daily(daily_28, 28, orders=orders, sku_barcode_map=sku_barcode_map)
    sales_60 = calc_sales_from_daily(daily_28, 60, orders=orders, sku_barcode_map=sku_barcode_map)

    # 复合 key 缓存(主循环每SKU×每仓×3窗口调用get_sales, 免重复拼串/查map——等价重构)
    _sales_key_cache = {}
    def get_sales(sales_dict, sku):
        """按 sku 查询日销，优先用复合 key sku|barcode，降级为 sku（key 缓存）"""
        key = _sales_key_cache.get(sku)
        if key is None:
            barcode = sku_barcode_map.get(sku, '')
            key = f"{sku}|{barcode}" if barcode else sku
            _sales_key_cache[sku] = key
        val = sales_dict.get(key)
        if val is not None:
            return val
        return sales_dict.get(sku, 0)

    def fused_ds(ds7, ds14, ds28):
        return rolling_predict(ds7, ds14, ds28)

    if mode == 'bbcc':
        c_lead = int(cfg.get('b_to_c_days', '0')) + int(cfg.get('c_safety_days', '0'))
        lead_time = c_lead
    else:
        lead_time = int(cfg.get('lead_time_days', '0'))

    season_key = f'season_config_{mode}'
    season_val = db.table('replenishment_config').select('*').eq('key', season_key).execute().data
    season_config = json.loads(season_val[0]['value']) if season_val and season_val[0].get('value') else []
    active_factor = 1.0
    for s in season_config:
        if isinstance(s, dict) and s.get("enabled") and float(s.get("factor", 1.0)) > active_factor:
            active_factor = float(s["factor"])

    suggestions = []

    if mode == 'bbcc':
        items = db.table("inventory").select("*").in_("warehouse_type", ["platform", "platform_b", "own"]).eq("channel", channel).execute().data
        agg = {}; wh_detail = {}; b_stock = {}; b_transit = {}; own_stock = {}
        for inv in items:
            sku = inv.get("sku", "")
            if sku not in agg:
                agg[sku] = {'available':0,'transit':0,'c_transit':0,'safety':0,'safety_days':0,'warehouses':set()}
                wh_detail[sku] = []; b_stock[sku] = 0; b_transit[sku] = 0; own_stock[sku] = 0
            wt = inv.get('warehouse_type',''); qty = int(inv.get("available_qty") or 0); tty = int(inv.get("in_transit_qty") or 0); ctt = int(inv.get("c_transit") or 0)
            if wt == 'platform_b':
                b_stock[sku] += qty          # B仓可用
                b_transit[sku] += tty        # 供应商→B仓在途
            elif wt == 'own': own_stock[sku] += qty  # 自有仓(集货/三方): B 维度补给源(补货时从自有调)
            else:
                agg[sku]['available'] += qty
                agg[sku]['transit'] += tty   # C仓供应商在途(传统口径, BBCC 不用)
                agg[sku]['c_transit'] += ctt  # B→C调拨在途: C 侧可用补充(用户链路: C可用+B-C调拨在途满足则不补)
            agg[sku]['safety'] += int(inv.get("safety_qty") or 0)
            sd = float(inv.get('safety_days') or 0)
            if sd > agg[sku]['safety_days']: agg[sku]['safety_days'] = sd
            wh_name = inv.get('warehouse','')
            if wh_name: agg[sku]['warehouses'].add(wh_name); wh_detail[sku].append({'warehouse':wh_name,'type':wt,'available':qty,'transit':tty})

        for sku, st in agg.items():
            avail = st['available']; transit = st['transit']; c_transit = st['c_transit']; safety = st['safety']
            ds7 = round(get_sales(sales_7, sku), 1)
            ds14 = round(get_sales(sales_14, sku), 1)
            ds28 = round(get_sales(sales_28, sku), 1)
            ds60 = round(get_sales(sales_60, sku), 1)
            sel_ds = round(fused_ds(ds7, ds14, ds28) * active_factor, 1)
            sku_safety_days = st['safety_days']
            safety_days = sku_safety_days if sku_safety_days > 0 else float(cfg.get('safety_multiplier', '0'))
            effective_safety = round(sel_ds * safety_days, 1) if sel_ds > 0 else 0
            # 业务链路(用户口径, SKU-0967 验证: 日销21.6×lead6=129.6−C可用17−c_transit0=112.6):
            # 1) C缺口 = 日销×lead_time − C可用 − B→C调拨在途(c_transit)
            #    注意: BBCC 的 C 仓在途只用 c_transit(B→C调拨在途), 不减 in_transit(供应商→C, 传统模式概念)
            #    保留1位小数(例112.6)——不得 round() 取整(曾取整致112.6→113, 偏离业务值)
            c_gap = max(round(sel_ds * lead_time - avail - c_transit, 1), 0) if sel_ds > 0 else 0
            # 2) C建议补 = C缺口(剩余缺口, 调拨到C即满足)——显示剩余缺口本身(例112.6), 箱规由前端/后续取整
            suggested = c_gap
            # 3) B 维度: B缺口 = C缺口 − (B可用 + 供应商→B在途)
            #    若 B可用+供应商到B ≥ C缺口 → B 建议补不显示值(0)
            b_available = b_stock.get(sku, 0)
            b_in_transit = b_transit.get(sku, 0)  # 供应商→B仓在途(B仓行 in_transit_qty)
            b_cover = b_available + b_in_transit
            b_gap = max(round(c_gap - b_cover, 1), 0) if c_gap > 0 else 0
            b_ship_days = int(cfg.get('ship_to_b_days', '0'))  # 自有→B调拨时间(规则页)
            # 4) B建议补 = B缺口 + 调拨期消耗(日销×(自有-b仓时间+安全天数)), 箱规取整
            #    例: 110.6 + 21.6×(3+3)=129.6 → 240.2 → 箱规24 → 264
            b_replenish = round(b_gap + sel_ds * b_ship_days + effective_safety, 1) if b_gap > 0 else 0
            prod = products.get(sku, {})
            box = int(prod.get('box_qty', 1) or 1)
            # C建议补 = C缺口剩余值(例112.6)本身, 不提前箱规取整; B建议补才箱规取整(例264)
            suggested = c_gap
            b_box_qty = (b_replenish + box - 1) // box * box if b_replenish > 0 else 0
            after_stock = avail + transit + suggested
            after_turnover = round(after_stock / sel_ds, 1) if sel_ds > 0 else 999
            days_to_empty = round(avail / sel_ds, 1) if sel_ds > 0 else 999
            combined_turnover_current = round((avail + transit + b_stock.get(sku, 0)) / sel_ds, 1) if sel_ds > 0 else None
            combined_turnover = round((avail + transit + suggested + b_stock.get(sku, 0) + b_box_qty) / sel_ds, 1) if sel_ds > 0 else None
            t7 = '📈' if ds7 > ds14 * 1.15 else ('📉' if ds7 < ds14 * 0.85 else '➡️')
            t14 = '📈' if ds14 > ds28 * 1.15 else ('📉' if ds14 < ds28 * 0.85 else '➡️')
            trend_text = f"近7{t7} 近14{t14}"
            if sel_ds > 0 and sel_ds < 5 and combined_turnover_current is not None and combined_turnover_current > 90:
                trend_text += " 销量极低，库存积压"
            elif ds7 == 0 and ds14 == 0 and ds28 > 0: trend_text += " 持续下行（近14天无销量）"
            elif ds7 > ds14 * 1.15 and ds14 > ds28 * 1.1: trend_text += " 持续上行"
            elif ds7 < ds14 * 0.85 and ds14 < ds28 * 0.9: trend_text += " 持续下行"
            elif ds7 > ds14 * 1.15: trend_text += " 7天抬头"
            elif ds7 < ds14 * 0.85: trend_text += " 7天走弱"
            else: trend_text += " 平稳"

            parts = []
            # ◆ 1. 趋势(优先级1: 方向)
            if sel_ds > 0: parts.append(trend_text)
            # ◆ 2. C建议补(优先级2: 需求端结论)
            if c_gap > 0:
                if b_gap <= 0:
                    parts.append(f"C建议补{suggested}件(缺口{c_gap},B仓可覆盖)")
                else:
                    parts.append(f"C建议补{suggested}件 · B建议补{b_box_qty}件(缺口{b_gap},调拨消耗{round(sel_ds*b_ship_days+effective_safety,1)},箱规{box})")
                if b_in_transit > 0 and b_gap > 0:
                    parts.append(f"B仓仅{b_available}件需从自有仓调(供应商到B在途{b_in_transit})")
                elif b_gap > 0:
                    parts.append(f"B仓仅{b_available}件需从自有仓调")
            # ◆ 3. 风险(优先级3: 仓储费/周转)
            if b_gap > 0:
                c_cover = round((avail + transit) / sel_ds, 1) if sel_ds > 0 else 0
                b_idle = max(round(c_cover - b_ship_days, 1), 0)
            else: b_idle = 0
            b_free = int(cfg.get('b_free_days', '15'))
            if b_idle > b_free: parts.append(f"🔴 超{b_free}天免费期有仓储费")
            elif b_idle > b_free - 5: parts.append(f"⚠️ 接近{b_free}天免费期")
            tw90_val = int(cfg.get('turnover_warning_90', '90'))
            has_replen = (suggested > 0 or b_box_qty > 0)
            turn_check = combined_turnover if has_replen and combined_turnover is not None else combined_turnover_current
            if turn_check is not None and turn_check > tw90_val:
                label = "补后综转" if has_replen else "当前综转"
                parts.append(f"🔴 {label}{turn_check}天超{tw90_val}天")
            elif turn_check is not None and turn_check > tw90_val - 15:
                label = "补后综转" if has_replen else "当前综转"
                parts.append(f"⚠️ {label}{turn_check}天接近{tw90_val}天")
            if sel_ds <= 0:
                if b_stock.get(sku, 0) > 0: parts.append("🔴 近30天无销量，B仓库存积压")
                elif avail > 0: parts.append("🔴 近30天无销量，C仓库存积压")
                else: parts.append("⚪ 近30天无销量")
            if b_gap > 0 and not any('B仓仅' in p for p in parts):
                parts.append(f"B仓仅{b_available}件需从自有仓调")
            if b_gap > 0:
                parts.append(f"B仓预计空闲{b_idle}天")
            if not parts: parts.append("库存充足")
            note = " · ".join(parts)
            c_turnover = round(avail / sel_ds, 1) if sel_ds > 0 else None
            transit_turnover = round(transit / sel_ds, 1) if sel_ds > 0 else None
            suggestions.append({
                "sku": sku, "barcode": sku_barcode_map.get(sku, ''), "product_name": prod.get('product_name', ''), "brand": prod.get('brand', ''),
                "store": prod.get('store', ''), "category": prod.get('category', ''),
                "available_qty": avail, "safety_qty": safety, "in_transit_qty": transit,
                "c_transit": c_transit, "b_transit": b_transit.get(sku, 0),  # B→C调拨在途/供应商→B在途(与进销存同源)
                "b_stock": b_stock.get(sku, 0), "c_stock": avail, "b_gap": b_gap,
                "daily_sales": sel_ds, "daily_sales_7": round(ds7, 1), "daily_sales_14": round(ds14, 1), "daily_sales_28": round(ds28, 1), "daily_sales_60": round(ds60, 1),
                "raw_suggested": c_gap, "suggested_qty": suggested,
                "b_suggested": b_box_qty, "b_replenish_raw": b_replenish,
                "days_to_empty": days_to_empty, "after_turnover": after_turnover,
                "c_turnover": c_turnover, "transit_turnover": transit_turnover,
                "combined_turnover_current": combined_turnover_current, "combined_turnover": combined_turnover,
                "note": note, "warehouse_detail": wh_detail.get(sku, []),
            })
    else:
        # 传统模式
        all_inv = db.table("inventory").select("*").eq("channel", channel).execute().data
        # 按仓库维度加载日销（每个仓库的日销独立，非 SKU 总日销）
        _wh_daily = {}
        _wh_wh_sales = {}
        _inv_wh = [x for x in db.table("inventory").select("*").in_("warehouse_type", ["platform"]).eq("channel", channel).execute().data]
        _wh_names = set(x.get('warehouse','') for x in _inv_wh)
        # 优化：一次性加载所有仓库的快照 + 当天订单，避免 7 次独立 SQL（PA 15s 代理超时）
        if _wh_names:
            try:
                from app.core.database import get_conn
                conn = get_conn()
                cutoff = (datetime.now(timezone.utc) - timedelta(days=28)).strftime('%Y-%m-%d')
                _wns = list(_wh_names)
                _ph = ','.join('?' * len(_wns))
                _snap_rows = conn.execute(
                    f"SELECT date, sku, warehouse, order_count FROM daily_sales_snapshot WHERE date>=? AND channel=? AND warehouse IN ({_ph})",
                    (cutoff, channel) + tuple(_wns)
                ).fetchall()
                _wh_raw = {}
                for _r in _snap_rows:
                    _wh = _r[2]; _sku = _r[1]; _key = f"{_sku}|{sku_barcode_map.get(_sku,'')}" if sku_barcode_map and sku_barcode_map.get(_sku) else _sku
                    _wh_raw.setdefault(_wh, {}).setdefault(_key, {})[_r[0]] = (_r[3] or 0)
                # 当天 orders 补充（原始 SQL 一次性）
                today = datetime.now(timezone.utc).strftime('%Y-%m-%d')
                _today_rows = conn.execute(
                    "SELECT sku, warehouse, quantity, ordered_at FROM orders WHERE ordered_at>=? AND channel=? AND (deleted_at='')",
                    (today, channel)
                ).fetchall()
                for _o in _today_rows:
                    _wh = _o[1] or ''; _sku = _o[0]; _key = f"{_sku}|{sku_barcode_map.get(_sku,'')}" if sku_barcode_map and sku_barcode_map.get(_sku) else _sku
                    _dt = str(_o[3] or '')[:10]
                    if _dt >= cutoff:
                        _d = _wh_raw.setdefault(_wh, {}).setdefault(_key, {})
                        _d[_dt] = _d.get(_dt, 0) + (_o[2] or 0)
                for _wn in _wh_names:
                    _daily = _wh_raw.get(_wn, {})
                    _wh_wh_sales[_wn] = {
                        '7': calc_sales_from_daily(_daily, 7),
                        '14': calc_sales_from_daily(_daily, 14),
                        '28': calc_sales_from_daily(_daily, 28),
                    }
            except Exception as _e:
                logger.warning(f"[replen] batch daily_sales fallback: {_e}")
                # 回退到逐仓加载
                for _wn in _wh_names:
                    _wh_daily_28 = load_daily_sales(28, db, sku_barcode_map=sku_barcode_map, channel=channel, warehouse=_wn)
                    _wh_wh_sales[_wn] = {
                        '7': calc_sales_from_daily(_wh_daily_28, 7),
                        '14': calc_sales_from_daily(_wh_daily_28, 14),
                        '28': calc_sales_from_daily(_wh_daily_28, 28),
                    }
        for inv in _inv_wh:
            sku = inv.get("sku", "")
            warehouse = inv.get("warehouse", "")
            avail = int(inv.get("available_qty") or 0)
            transit = int(inv.get("in_transit_qty") or 0)
            safety = int(inv.get("safety_qty") or 0)
            _wh_s = _wh_wh_sales.get(warehouse, {})
            ds7 = round(get_sales(_wh_s.get('7', {}), sku), 1)
            ds14 = round(get_sales(_wh_s.get('14', {}), sku), 1)
            ds28 = round(get_sales(_wh_s.get('28', {}), sku), 1)
            ds60 = round(get_sales(sales_60, sku), 1)
            sel_ds = round(fused_ds(ds7, ds14, ds28) * active_factor, 1)
            safety_days = float(cfg.get('safety_multiplier', '0'))
            effective_safety = round(sel_ds * safety_days) if sel_ds > 0 else 0
            suggested = max(round(sel_ds * lead_time + effective_safety - avail - transit), 0) if sel_ds > 0 else 0

            prod = products.get(sku, {})
            box = int(prod.get('box_qty', 1) or 1)
            box_qty = (suggested + box - 1) // box * box if suggested > 0 else 0
            after_stock = avail + transit + box_qty
            after_turnover = round(after_stock / sel_ds, 1) if sel_ds > 0 else 999
            days_to_empty = round(avail / sel_ds, 1) if sel_ds > 0 else 999
            t7 = '📈' if ds7 > ds14 * 1.15 else ('📉' if ds7 < ds14 * 0.85 else '➡️')
            t14 = '📈' if ds14 > ds28 * 1.15 else ('📉' if ds14 < ds28 * 0.85 else '➡️')
            trend_text = f"近7{t7} 近14{t14}"
            tw90 = int(cfg.get('turnover_warning_90', '90'))
            parts = []
            if box_qty > 0:
                if after_turnover > tw90: parts.append(f"🔴 补后周转{after_turnover}天超{tw90}天")
                elif after_turnover > tw90 - 15: parts.append(f"⚠️ 补后周转{after_turnover}天接近{tw90}天")
            else:
                if days_to_empty > tw90: parts.append(f"🔴 当前周转{days_to_empty}天超{tw90}天")
                elif days_to_empty > tw90 - 15: parts.append(f"⚠️ 当前周转{days_to_empty}天接近{tw90}天")
            if sel_ds <= 0:
                if avail > 0: parts.append("🔴 近30天无销量，库存积压")
                else: parts.append("⚪ 近30天无销量")
            if sel_ds > 0: parts.append(trend_text)
            # 箱数提示 + 补1箱智能提醒
            if box_qty > 0:
                _boxes = box_qty // box
                if _boxes > 1:
                    parts.append(f"建议补{box_qty}件（{_boxes}箱）· 箱规{box}件")
                else:
                    parts.append(f"建议补{box_qty}件（1箱）· 箱规{box}件")
                    # 补1箱时：检查其他仓是否有高库存且低日销的 SKU
                    _cross_wh = []
                    for _x in all_inv:
                        if _x.get('sku')==sku and _x.get('warehouse')!=warehouse:
                            _x_avail = int(_x.get('available_qty') or 0)
                            if _x_avail > avail * 2:  # 其他仓库存远大于当前仓
                                _cross_wh.append(f"{_x.get('warehouse','')}还有{_x_avail}件")
                    if _cross_wh:
                        parts.append(f"跨仓调拨: {', '.join(_cross_wh[:2])}，建议人工复核是否需补货")
                    else:
                        parts.append("仅补1箱，建议人工复核")
            else:
                wh_data = [x for x in all_inv if x.get('sku')==sku and x.get('warehouse')!=warehouse]
                if wh_data: parts.append(f"跨仓提示: {wh_data[0].get('warehouse','')}还有{wh_data[0].get('available_qty',0)}件")
            if not parts: parts.append("库存充足")
            note = " · ".join(parts)
            suggestions.append({
                "sku": sku, "barcode": sku_barcode_map.get(sku, ''), "product_name": prod.get('product_name', ''), "brand": prod.get('brand', ''),
                "store": prod.get('store', ''), "warehouse": warehouse, "category": prod.get('category', ''),
                "available_qty": avail, "safety_qty": safety, "in_transit_qty": transit,
                "daily_sales": sel_ds, "daily_sales_7": round(ds7, 1), "daily_sales_14": round(ds14, 1), "daily_sales_28": round(ds28, 1), "daily_sales_60": round(ds60, 1),
                "suggested_qty": box_qty, "after_turnover": after_turnover, "days_to_empty": days_to_empty,
                "note": note,
            })

    # 排序：需要补货的排最前（suggested_qty>0 或 b_suggested>0），
    # 其次缺口越大越优先，再按日销降序，最后 SKU 稳定排序
    # （避免 6900+ 条"库存充足"排在前面，首屏全是建议补 - 被误解为无数据）
    def _needs(s):
        return 1 if (s.get('suggested_qty') or 0) > 0 or (s.get('b_suggested') or 0) > 0 else 0
    suggestions.sort(key=lambda s: (
        -_needs(s),
        # 在途>0 优先(需补货但已有在途的先看到, 避免前1000条全在途0误判为bug)
        -(_needs(s) * (int(s.get('in_transit_qty') or 0) > 0) * 2),
        -(s.get('suggested_qty') or 0),
        -(s.get('b_suggested') or 0),
        -(s.get('daily_sales') or 0),
        s.get('sku', ''),
    ))

    # 搜索过滤(排序后, 搜索词相关性优先展示)
    if search:
        _sq = search.lower()
        suggestions = [s for s in suggestions if _sq in str(s.get('sku','')).lower() or _sq in str(s.get('product_name','')).lower() or _sq in str(s.get('barcode','')).lower()]
    # 写入缓存
    try:
        from app.core.replenishment_cache import set_cache
        set_cache(mode, channel, days, {"data": suggestions}, db)
    except Exception as e:
        logger.warning(f"write replenishment cache: {e}")

    # 生成/关闭补货告警（按渠道隔离，只管理自己 source='replenishment_engine' 的告警，
    # 不干扰规则引擎 source='rules_engine' 生成的紧急补货告警）
    try:
        from app.core.database import get_conn
        conn = get_conn()
        active = {r[0] for r in conn.execute(
            "SELECT related_sku FROM alerts WHERE alert_type='replenish' AND status='active' AND channel=? AND source='replenishment_engine'",
            (channel,)).fetchall()}
        for s in suggestions:
            sku = s.get('sku', '')
            qty = s.get('suggested_qty', 0) or s.get('b_suggested', 0) or 0
            if qty > 0 and sku not in active:
                # 补货告警 warehouse_type: 该 SKU 库存主体(优先有货仓, 与看板 B/C/自有 分布口径一致)
                wt = ''
                try:
                    _r = conn.execute("SELECT warehouse_type FROM inventory WHERE sku=? AND channel=? ORDER BY CASE WHEN available_qty>0 THEN 0 ELSE 1 END, available_qty DESC LIMIT 1", (sku, channel)).fetchone()
                    wt = _r[0] if _r else ''
                except Exception: pass
                try:
                    conn.execute("INSERT INTO alerts(alert_type,title,description,severity,source,related_sku,status,channel,warehouse_type) VALUES(?,?,?,?,?,?,?,?,?)",
                        ("replenish", f"需补货: {s.get('product_name', sku)}",
                         f"建议补{qty}件, 可撑{s.get('days_to_empty', 0)}天",
                         "warning", "replenishment_engine", sku, "active", channel, wt))
                except Exception as e: logger.warning(f"[replenish] insert alert: {e}")
            elif qty == 0 and sku in active:
                try:
                    conn.execute("UPDATE alerts SET status='closed' WHERE alert_type='replenish' AND related_sku=? AND channel=? AND status='active' AND source='replenishment_engine'", (sku, channel))
                except Exception as e: logger.warning(f"[replenish] close alert: {e}")
        conn.commit()
    except Exception as e:
        logger.warning(f"[replenish] batch alerts: {e}")

    if page > 0 and page_size > 0:
        _total = len(suggestions)
        return ok({"items": suggestions[(page - 1) * page_size: page * page_size], "total": _total, "page": page, "page_size": page_size})
    return ok(suggestions)

@router.get('/replenishment/compare')
def compare_replenishment_sources(days: int = 28, db = get_db()):
    """对比不同数据源的补货建议"""
    results = {}
    for src in ['', 'jd_purchase', 'cleansing']:
        try:
            data = get_replenishment_suggestions(days=days, source=src, mode='bbcc', db=db)
            items = data.get("data") if isinstance(data, dict) and "data" in data else data
            results[src or 'all'] = items[:5] if items else []
        except: results[src or 'all'] = []
    return ok(results)


@router.get('/export-orders')
def export_orders_excel(channel: str = 'jd', db = get_db()):
    from openpyxl import Workbook
    from io import BytesIO; from fastapi.responses import Response; from urllib.parse import quote
    orders = [o for o in db.table("orders").select("*").eq("channel", channel).order("id", desc=True).limit(2000).execute().data if not (o.get("deleted_at") or "")]
    wb = Workbook(); ws = wb.active; ws.title = "订单"
    ws.append(["下单日期","订单号","69码","店铺","仓库","商品","金额","状态","入库日期","平台"])
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    hf = PatternFill(start_color="1d4ed8",end_color="1d4ed8",fill_type="solid")
    hfn = Font(bold=True,color="ffffff",size=11)
    thin = Border(left=Side(style='thin',color='e2e8f0'),right=Side(style='thin',color='e2e8f0'),top=Side(style='thin',color='e2e8f0'),bottom=Side(style='thin',color='e2e8f0'))
    for c in ws[1]: c.fill=hf; c.font=hfn; c.alignment=Alignment(horizontal='center'); c.border=thin
    for o in orders:
        ws.append([str(o.get('ordered_at',''))[:10],o.get('order_no',''),o.get('barcode',''),o.get('store',''),
            o.get('warehouse',''),o.get('product_name',''),o.get('total_amount',0),o.get('order_status',''),
            str(o.get('paid_at',''))[:10],o.get('platform','')])
        for c in ws[ws.max_row]: c.border=thin; c.alignment=Alignment(horizontal='center')
    buf = BytesIO(); wb.save(buf); buf.seek(0)
    return Response(content=buf.getvalue(),media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition":f"attachment; filename*=UTF-8''orders_{datetime.now(timezone.utc).strftime('%Y%m%d')}.xlsx"})


@router.get('/export-inventory')
def export_inventory_excel(channel: str = 'jd', wh_type: str = '', db = get_db()):
    from openpyxl import Workbook
    from io import BytesIO; from fastapi.responses import Response; from urllib.parse import quote
    query = db.table("inventory").select("*").eq("channel", channel)
    if wh_type:
        query = query.eq("warehouse_type", wh_type)
    inv = query.execute().data
    wb = Workbook(); ws = wb.active; ws.title = "库存"
    ws.append(["SKU","69码","商品","仓库","仓库类型","渠道","可用","在途","安全库存","期初库存","当月采购入库","当月出库","在库周转"])
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    hf = PatternFill(start_color="1d4ed8",end_color="1d4ed8",fill_type="solid")
    hfn = Font(bold=True,color="ffffff",size=11)
    thin = Border(left=Side(style='thin',color='e2e8f0'),right=Side(style='thin',color='e2e8f0'),top=Side(style='thin',color='e2e8f0'),bottom=Side(style='thin',color='e2e8f0'))
    for c in ws[1]: c.fill=hf; c.font=hfn; c.alignment=Alignment(horizontal='center'); c.border=thin
    for i in inv:
        td = round(i.get('turnover_days',0) or 0, 1) if (i.get('turnover_days') or 0) > 0 else None
        ws.append([i.get('sku',''),i.get('barcode',''),i.get('product_name',''),i.get('warehouse',''),i.get('warehouse_type',''),
            i.get('channel',''),i.get('available_qty',0),i.get('in_transit_qty',0),i.get('safety_qty',0),
            i.get('beginning_stock',0),i.get('month_inbound',0),i.get('month_outbound',0),td])
        for c in ws[ws.max_row]: c.border=thin; c.alignment=Alignment(horizontal='center')
    buf = BytesIO(); wb.save(buf); buf.seek(0)
    return Response(content=buf.getvalue(),media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition":f"attachment; filename*=UTF-8''inventory_{datetime.now(timezone.utc).strftime('%Y%m%d')}.xlsx"})


@router.get('/export-purchase')
def export_purchase_excel(days: int = 28, mode: str = 'bbcc', channel: str = 'jd', db = get_db()):
    """导出补货建议为 Excel"""
    from openpyxl import Workbook
    from io import BytesIO; from fastapi.responses import Response; from urllib.parse import quote
    from datetime import timedelta, timezone

    data = get_replenishment_suggestions(days=days, source='', mode=mode, channel=channel, db=db)
    replen = data.get("data") if isinstance(data, dict) and "data" in data else data

    wb = Workbook(); ws = wb.active; ws.title = "补货建议"
    if mode == 'bbcc':
        headers = ["序号","SKU","69码","商品","仓库","B仓可用库存","B仓周转","C仓总和可用","B-C调拨在途",
            "日销(融合/7/14/28)","C仓周转","B→C调拨周转","C仓建议补","B仓需补","当前综转","补后综转","备注"]
    else:
        headers = ["序号","SKU","69码","商品","仓库","现有","在途","日销(融合/7/14/28)","安全线","在库周转","补后周转","建议补","备注"]
    ws.append(headers)

    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    hf = PatternFill(start_color="1d4ed8",end_color="1d4ed8",fill_type="solid")
    hfn = Font(bold=True,color="ffffff",size=11)
    thin = Border(left=Side(style='thin',color='e2e8f0'),right=Side(style='thin',color='e2e8f0'),top=Side(style='thin',color='e2e8f0'),bottom=Side(style='thin',color='e2e8f0'))
    for c in ws[1]: c.fill=hf; c.font=hfn; c.alignment=Alignment(horizontal='center'); c.border=thin

    for i, r in enumerate(replen, 1):
        if mode == 'bbcc':
            b_turn = f"{round(r['b_stock']/r['daily_sales'],1)}天" if r.get("b_stock",0)>0 and r.get("daily_sales",0)>0 else ""
            ws.append([i,r["sku"],r.get("barcode","-"),r["product_name"],"B仓",r.get("b_stock","-"),b_turn,r.get("c_stock",r.get("available_qty",0)),
                r.get("in_transit_qty",0),r.get("daily_sales",0),r.get("daily_sales_7",0),r.get("daily_sales_14",0),r.get("daily_sales_28",0),
                f"{r.get('c_turnover','∞')}天" if r.get('c_turnover') else "∞",f"{r.get('transit_turnover','∞')}天" if r.get('transit_turnover') else "∞",
                r.get("suggested_qty","-"),r.get("b_suggested","-"),
                f"{r.get('combined_turnover_current','∞')}天" if r.get('combined_turnover_current') else "∞",
                f"{r.get('combined_turnover','∞')}天" if r.get('combined_turnover') else "∞",r.get("note","")])
        else:
            after_turn = f"{r['after_turnover']}天" if r.get("suggested_qty",0)>0 and r.get("after_turnover") else ""
            ws.append([i,r["sku"],r.get("barcode","-"),r["product_name"],r.get("warehouse",r.get("store","-")),r.get("available_qty",0),r.get("in_transit_qty",0),
                r.get("daily_sales",0),r.get("daily_sales_7",0),r.get("daily_sales_14",0),r.get("daily_sales_28",0),
                r.get("safety_qty",0),f"{r.get('days_to_empty','∞')}天" if r.get('days_to_empty',999)<999 else "∞",
                after_turn,r.get("suggested_qty","-"),r.get("note","")])
        for c in ws[ws.max_row]: c.border=thin; c.alignment=Alignment(horizontal='center')

    widths = [6,14,22,10]+[12]*(len(headers)-4)
    for i, w in enumerate(widths, 1):
        if i <= len(headers): ws.column_dimensions[ws.cell(1,i).column_letter].width = w

    ws2 = wb.create_sheet("汇总")
    ws2.append(["补货建议汇总"]); ws2.append(["生成时间",datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M")])
    ws2.append(["模式","BBCC送仓" if mode=='bbcc' else "传统多仓"])
    ws2.append(["SKU数",len(replen)])
    ws2.merge_cells('A1:D1'); ws2['A1'].font = Font(bold=True, size=14)

    buf = BytesIO(); wb.save(buf); buf.seek(0)
    filename = f"补货建议_{mode}_{datetime.now(timezone.utc).strftime('%Y%m%d')}.xlsx"
    return Response(content=buf.getvalue(),media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition":f"attachment; filename*=UTF-8''{quote(filename)}"})
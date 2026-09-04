from fastapi import APIRouter, Depends, UploadFile, File, Form, HTTPException
from datetime import datetime, timezone
import json, csv, io, re, os, uuid
from openpyxl import load_workbook
import logging
logger = logging.getLogger(__name__)
from app.core.database import get_db, get_conn, submit_task, get_task, backup_db
from app.core.response import ok, fail
from app.core.cleansing_parser import parse_file, cleanse_value
from app.core.cleansing_templates import load_custom_fields, save_custom_fields, list_templates, save_template, delete_template, get_system_fields
from app.api.routes.ws import broadcast
from app.api.routes.insights import auto_adjust_inventory

router = APIRouter(prefix="/api/cleansing", tags=["cleansing"])

# ─── 系统目标字段定义 ────────────────────────────────────────────────────────

# ─── 自定义字段存储 ────────────────────────────────────────────────────────────

CUSTOM_FIELDS_PATH = '/home/Overtrees/Supplykit/backend/custom_fields.json'

def parse_file(content, filename):
    if filename.lower().endswith('.csv'):
        text = content.decode('utf-8-sig', errors='ignore')
        return list(csv.DictReader(io.StringIO(text)))
    wb = load_workbook(io.BytesIO(content), data_only=True)
    ws = wb[wb.sheetnames[0]]
    raw = list(ws.iter_rows(values_only=True))
    if not raw:
        return ok([])
    headers = [str(c).strip() if c is not None else '' for c in raw[0]]
    return [{headers[i]: raw[r][i] for i in range(len(headers))} for r in range(1, len(raw))]

# ─── 检测接口 ────────────────────────────────────────────────────────────────

@router.post('/detect')
async def detect_columns(file: UploadFile = File(...)):
    content = await file.read()
    rows = parse_file(content, file.filename)
    if not rows:
        return {'ok': False, 'error': '文件为空'}
    cols = []
    for key in rows[0].keys():
        samples = []
        for r in rows[:5]:
            v = r.get(key)
            if v is not None and str(v).strip():
                samples.append(str(v)[:60])
        cols.append({'name': key, 'samples': samples[:3], 'count': len(rows)})
    return {'ok': True, 'columns': cols, 'total': len(rows), 'file': file.filename}

# ─── 预览接口 ────────────────────────────────────────────────────────────────

@router.post('/preview')
async def preview_cleansing(file: UploadFile = File(...), mapping: str = Form('')):
    content = await file.read()
    rows = parse_file(content, file.filename)
    if not rows:
        return {'ok': False, 'error': '文件为空'}
    try:
        mapping_config = json.loads(mapping) if mapping else {}
    except json.JSONDecodeError:
        return {'ok': False, 'error': '映射配置格式错误'}

    # 检测 warehouse 字段一致性
    warehouse_in_file = set()
    warehouse_target = None
    for src_col, cfg in mapping_config.items():
        if cfg.get('target') == 'warehouse':
            warehouse_target = src_col
            break
    warnings = []
    # 订单导入若未映射仓库列: 订单归"未知"仓, 日销无法按 C 仓/B 仓归因(影响补货口径)
    _is_order = any(cfg.get('target') == 'order_no' for cfg in mapping_config.values())
    if _is_order and not warehouse_target:
        warnings.append('⚠️ 订单未映射「仓库」列——订单将归入"未知"仓，日销无法按 C 仓/B 仓归因（影响补货/濒临断货口径），建议映射仓库列')
    if warehouse_target:
        for row in rows[:50]:
            w = str(row.get(warehouse_target, '')).strip()
            if w:
                warehouse_in_file.add(w)
        # 查现有库存中的仓库
        db = get_db()
        existing = set()
        for i in db.table("inventory").select("warehouse").execute().data or []:
            w = (i.get('warehouse') or '').strip()
            if w:
                existing.add(w)
        unknown = warehouse_in_file - existing
        if unknown:
            warnings.append(f'以下仓库名在库存表中不存在，订单页平台库存将显示"—"：{", ".join(sorted(unknown))}')

    preview_rows = []
    for row in rows[:50]:
        result = {'_source': {}}
        for src_col, cfg in mapping_config.items():
            target = cfg.get('target', '')
            if not target:
                continue
            raw_val = row.get(src_col, '')
            cleaned = cleanse_value(raw_val, cfg)
            result['_source'][src_col] = str(raw_val)[:80] if raw_val is not None else ''
            result[target] = cleaned
        preview_rows.append(result)

    return {'ok': True, 'preview': preview_rows, 'total': len(rows), 'mapped': len(mapping_config), 'warnings': warnings}

# ─── 执行清洗 ────────────────────────────────────────────────────────────────

def _run_cleansing(content: bytes, filename: str, mapping_json: str, target: str, template_name: str = '', channel: str = 'jd', conflict_mode: str = 'overwrite'):
    """清洗核心逻辑，含格式校验 → 业务校验 → 补全推断"""
    db = get_db()
    task_id = f"clean_{datetime.now(timezone.utc).strftime('%H%M%S')}"
    # 强制恢复 WAL 模式（之前可能因配额满降级为 DELETE，大批量写入极慢）
    try:
        conn = get_conn()
        conn.execute("PRAGMA journal_mode=WAL")
    except Exception:
        pass
    rows = parse_file(content, filename)
    if not rows:
        from app.core.database import update_task
        update_task(task_id, progress=100)
        return {'ok': False, 'error': '文件为空', 'success': 0, 'failed': 0, 'file': filename}
    try:
        mapping_config = json.loads(mapping_json) if mapping_json else {}
    except json.JSONDecodeError:
        from app.core.database import update_task
        update_task(task_id, progress=100)
        return {'ok': False, 'error': '映射配置格式错误', 'success': 0, 'failed': 0, 'file': filename}
    # 从映射中提取 data_source（前端通过自定义字段传入，或默认空）
    data_source = mapping_config.get('_meta', {}).get('data_source', '')

    # 加载用于校验和推断的参考数据
    products_map = {p["sku"]: p for p in db.table("products").select("*").eq("deleted_at", "").execute().data}
    inventory_map = {i["sku"]: i for i in db.table("inventory").select("*").execute().data}
    task_id = f"clean_{datetime.now(timezone.utc).strftime('%H%M%S')}"
    errors = []
    success = 0
    failed = 0
    is_inv = (target == 'inventory' or target == 'platform_inv' or target == 'inventory_b')
    is_inbound = (target == 'inbound')
    is_outbound = (target == 'outbound')
    is_product = (target == 'product')
    is_supplier = (target == 'supplier')
    platform_inv = (target == 'platform_inv')
    b_inv = (target == 'inventory_b')
    orders_to_insert = [] if not is_inv else None
    inv_to_insert = [] if is_inv else None
    batch_to_insert = []  # 库存导入带生产/截止日期 → 按批次写入
    inbound_to_insert = [] if is_inbound else None
    outbound_to_insert = [] if is_outbound else None
    product_to_insert = [] if is_product else None
    supplier_to_insert = [] if is_supplier else None
    sku_seen = set()

    for idx, row in enumerate(rows):
        row_errors = []
        data = {}

        # 更新进度 + WS 实时推送(切页也能收)
        if idx % 50 == 0:
            try:
                from app.core.database import update_task
                _p = round(idx / len(rows) * 100)
                update_task(task_id, progress=_p)
                try:
                    from app.api.routes.ws import broadcast_sync
                    broadcast_sync('cleansing_progress', {'task_id': task_id, 'progress': _p, 'status': 'running'})
                except Exception:
                    pass
            except Exception:
                pass

        # ─── 格式校验 + 字段映射 ──────────────────────────────────────
        for src_col, cfg in mapping_config.items():
            target_field = cfg.get('target', '')
            if not target_field: continue
            raw_val = row.get(src_col, '')
            try:
                cleaned = cleanse_value(raw_val, cfg)
                data[target_field] = cleaned
            except Exception as e:
                row_errors.append({'error_type': 'format_error', 'field_name': src_col,
                                   'raw_value': str(raw_val)[:100], 'error_message': str(e)[:100]})

        # ─── 补全推断 ──────────────────────────────────────────────
        sku = str(data.get('sku', ''))
        if sku and not data.get('product_name'):
            p = products_map.get(sku)
            if p:
                data['product_name'] = p.get('product_name', '')

        # ─── 业务校验 ──────────────────────────────────────────────
        if not data.get('ordered_at'):
            data['ordered_at'] = datetime.now(timezone.utc).strftime('%Y-%m-%d')

        # 去重处理（库存按 sku+仓库+批次(prod_date+exp_date) 复合去重）
        if is_inv:
            sku_val = sku or f"AUTO-{datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S')}-{success}"
            _wh_key = str(data.get('warehouse', '')) or ('平台仓' if platform_inv else 'B仓' if b_inv else '自有仓')
            _pd_key = str(data.get('prod_date', ''))[:10]
            _ed_key = str(data.get('exp_date', ''))[:10]
            _batch_key = (sku_val, _wh_key, _pd_key, _ed_key)
            if _batch_key in sku_seen:
                failed += 1; continue
            sku_seen.add(_batch_key)
            data['sku'] = sku_val
        else:
            order_no = data.get('order_no', '')
            if not order_no:
                order_no = f"AUTO-{datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S')}-{success}"
                data['order_no'] = order_no
            # 检查数据库中是否已存在（同一采购单号+同 SKU 才算重复）
            sku_key = data.get('sku', '')
            dup_row = db.table("orders").select("id").eq("order_no", order_no).eq("sku", sku_key).execute().data
            # 软删除订单不算重复（用户删单后重导同单应放行）
            if dup_row:
                dup_row = [x for x in dup_row if not (x.get("deleted_at") or "")]
            if dup_row:
                errors.append({'error_type': 'duplicate_order', 'field_name': 'order_no',
                               'raw_value': order_no, 'error_message': f'单号+SKU已存在: {order_no}/{sku_key}'})
                try:
                    if not db.table("quality_logs").select("id").eq("log_type","duplicate_order").eq("message",f'订单{order_no}商品{sku_key}已存在，跳过重复'[:100]).eq("source","cleansing").execute().data:
                        db.table("quality_logs").insert({"log_type":"duplicate_order","level":"warning",
                                "field_name":"order_no","message":f'订单{order_no}商品{sku_key}已存在，跳过重复',"source":"cleansing"}).execute()
                except Exception as e: print(f"[Cleansing] {e}")
                failed += 1; continue

        # 记录行错误（不影响继续处理，只是标记）
        for e in row_errors:
            errors.append(e)
            try:
                db.table("cleansing_errors").insert({
                    "task_id": task_id, "row_index": idx, "source_file": filename,
                    "error_type": e['error_type'], "field_name": e['field_name'],
                    "raw_value": e['raw_value'], "error_message": e['error_message'],
                    "raw_data": json.dumps(row, ensure_ascii=False, default=str),
                }).execute()
                # 同步写入 quality_logs，让异常页可查（跳过已存在的）
                if not db.table("quality_logs").select("id").eq("log_type",e['error_type']).eq("message",e['error_message'][:100]).eq("source","cleansing").execute().data:
                    db.table("quality_logs").insert({
                        "log_type": e['error_type'], "level": "warning",
                        "field_name": e['field_name'], "message": e['error_message'],
                        "source": "cleansing",
                    }).execute()
            except Exception as e: print(f"[Cleansing] {e}")

        # ─── 写入目标 ──────────────────────────────────────────────
        if is_inv:
            _wh_name = (str(data.get('warehouse', '')) or ('平台仓' if platform_inv else 'B仓' if b_inv else '自有仓'))[:100]
            _wh_type = 'platform_b' if b_inv else ('platform' if platform_inv else 'own')
            inv_to_insert.append({
                "sku": str(data.get('sku', ''))[:100],
                "product_name": str(data.get('product_name', ''))[:200],
                "store": str(data.get('store', '未知'))[:100],
                "warehouse": _wh_name,
                "warehouse_type": _wh_type,
                "warehouse_type": 'platform_b' if b_inv else ('platform' if platform_inv else 'own'),
                "channel": channel,
                "available_qty": int(float(data.get('available_qty', 0))),
                "locked_qty": int(float(data.get('locked_qty', 0))),
                "in_transit_qty": int(float(data.get('in_transit_qty', 0))),
                "c_transit": int(float(data.get('c_transit', 0) or 0)),
                "safety_qty": int(float(data.get('safety_qty', 0))),
            })
            # 若行含生产/截止日期 → 收集批次
            _pd = str(data.get('prod_date', ''))[:10]
            _ed = str(data.get('exp_date', ''))[:10]
            if _pd or _ed:
                _wh = (str(data.get('warehouse', '')) or ('平台仓' if platform_inv else 'B仓' if b_inv else '自有仓'))[:100]
                batch_to_insert.append((str(data.get('sku',''))[:100], _wh, 'platform_b' if b_inv else ('platform' if platform_inv else 'own'), channel, _pd, _ed, int(float(data.get('available_qty', 0)))))
            success += 1
        elif is_inbound:
            if not data.get('warehouse'):
                row_errors.append({'error_type': 'required_field', 'field_name': 'warehouse', 'raw_value': '', 'error_message': '入库记录必须映射仓库列'})
            inbound_to_insert.append({
                "sku": sku[:100],
                "product_name": str(data.get('product_name', ''))[:200],
                "quantity": int(float(data.get('quantity', 0))),
                "supplier": str(data.get('supplier', ''))[:100],
                "inbound_date": str(data.get('inbound_date', ''))[:50],
                "prod_date": str(data.get('prod_date', ''))[:10],
                "exp_date": str(data.get('exp_date', ''))[:10],
                "warehouse": str(data.get('warehouse', ''))[:100],
            })
            success += 1
        elif is_outbound:
            if not data.get('warehouse'):
                row_errors.append({'error_type': 'required_field', 'field_name': 'warehouse', 'raw_value': '', 'error_message': '出库记录必须映射仓库列'})
            outbound_to_insert.append({
                "sku": sku[:100],
                "product_name": str(data.get('product_name', ''))[:200],
                "quantity": int(float(data.get('quantity', 0))),
                "target_warehouse": str(data.get('target_warehouse', ''))[:100],
                "outbound_date": str(data.get('outbound_date', ''))[:50],
                "prod_date": str(data.get('prod_date', ''))[:10],
                "exp_date": str(data.get('exp_date', ''))[:10],
                "warehouse": str(data.get('warehouse', ''))[:100],
            })
            success += 1
        # 供应商导入
        elif is_supplier:
            _sup_code = str(data.get('supplier_code', '') or '')[:100]
            if not _sup_code:
                row_errors.append({'error_type': 'required_field', 'field_name': 'supplier_code', 'raw_value': data.get('supplier_code', ''), 'error_message': '供应商编号必填'})
            supplier_to_insert.append({
                "supplier_code": _sup_code,
                "supplier_name": str(data.get('supplier_name', ''))[:200],
                "contact_person": str(data.get('contact_person', ''))[:100],
                "contact_phone": str(data.get('contact_phone', ''))[:100],
                "score": int(float(data.get('score', 0) or 0)),
                "status": str(data.get('status', 'active'))[:50],
                "brand": str(data.get('brand', ''))[:200],
                "channel": channel,
            })
            success += 1
        # 商品信息导入
        if is_product:
            product_to_insert.append({
                "sku": sku[:100],
                "product_name": str(data.get('product_name', ''))[:200],
                "store": str(data.get('store', ''))[:100],
                "category": str(data.get('category', ''))[:100],
                "price": float(data.get('unit_price', data.get('price', 0))),
                "box_qty": int(float(data.get('box_qty', 0))),
                "barcode": str(data.get('barcode', ''))[:100],
                "brand": str(data.get('brand', ''))[:100],
                "unit": str(data.get('unit', ''))[:50],
                "status": str(data.get('status', 'active'))[:50],
                "weight": float(data.get('weight', 0)),
                "volume": float(data.get('volume', 0)),
                "best_before": str(data.get('best_before', ''))[:50],
            })
            success += 1
        elif not is_inv and not is_inbound and not is_outbound:
            orders_to_insert.append({
                "order_no": order_no, "store": str(data.get('store', '未知'))[:100],
                "sku": sku[:100],
                "product_name": str(data.get('product_name', ''))[:200],
                "quantity": int(float(data.get('quantity', 0))),
                "unit_price": float(data.get('unit_price', 0)),
                "total_amount": float(data.get('total_amount', 0)),
                "order_status": str(data.get('order_status', '已完成'))[:50],
                "ordered_at": str(data.get('ordered_at', ''))[:50],
                "paid_at": str(data.get('paid_at', ''))[:50],
                "data_source": data_source,
                "channel": channel,
                "barcode": str(data.get('barcode', ''))[:100],
                # GMV 金额明细(方案A): 缺省 0 平滑; GMV = total - discount + freight + tax
                "freight_amount": float(data.get('freight_amount', 0) or 0),
                "subsidy_amount": float(data.get('subsidy_amount', 0) or 0),
                "tax_amount": float(data.get('tax_amount', 0) or 0),
                "discount_amount": float(data.get('discount_amount', 0) or 0),
                "actual_amount": float(data.get('actual_amount', data.get('total_amount', 0)) or 0),
            })
            success += 1

    insert_table = 'products' if is_product else ('suppliers' if is_supplier else ('inventory' if is_inv else ('inbound_records' if is_inbound else ('outbound_records' if is_outbound else 'orders'))))
    data_list = product_to_insert if is_product else (supplier_to_insert if is_supplier else (inv_to_insert if is_inv else (inbound_to_insert if is_inbound else (outbound_to_insert if is_outbound else orders_to_insert))))
    if data_list:
        try:
            # 批量 upsert（executemany + ON CONFLICT + 分批 commit 防大事务 I/O error）
            if data_list:
                cols = list(data_list[0].keys())
                col_names = ", ".join(f'"{c}"' for c in cols)
                placeholders = ", ".join(["?"] * len(cols))
                # 构造 ON CONFLICT 更新子句（排除主键列）
                update_set = ", ".join([f'"{c}" = excluded."{c}"' for c in cols if c not in ('id', 'deleted_at')])
                if is_inv:
                    conflict_col = 'sku, warehouse, channel'  # 库存按 SKU+仓库+渠道 去重
                elif is_product:
                    conflict_col = 'sku'  # 商品按 SKU 去重（sku 有 UNIQUE 约束）
                elif is_supplier:
                    conflict_col = 'supplier_code'  # 供应商按编号去重（唯一约束）
                elif is_inbound:
                    conflict_col = 'sku, warehouse, channel, prod_date, exp_date, inbound_date'
                elif is_outbound:
                    conflict_col = 'sku, warehouse, channel, prod_date, exp_date, outbound_date'
                else:
                    conflict_col = 'order_no, sku'
                sql = f'INSERT INTO "{insert_table}" ({col_names}) VALUES ({placeholders}) ON CONFLICT({conflict_col}) DO UPDATE SET {update_set}'
                conn = get_conn()
                params_list = [[row.get(c) for c in cols] for row in data_list]
                # 分批写入（每 5000 条一次 commit，避免单事务过大触发 WAL I/O error）
                BATCH = 5000
                for i in range(0, len(params_list), BATCH):
                    conn.executemany(sql, params_list[i:i+BATCH])
                    conn.commit()
                # 入库/出库导入时自动累加 inventory 月汇总(确保主行=展开行之和)
                if is_inbound:
                    for _r in data_list:
                        try:
                            _sk = _r.get('sku',''); _wh = _r.get('warehouse',''); _qt = int(_r.get('quantity',0) or 0)
                            if _sk and _wh and _qt>0:
                                conn.execute("UPDATE inventory SET month_inbound = COALESCE(month_inbound,0)+? WHERE sku=? AND warehouse=? AND channel=? AND warehouse_type='own'", (_qt, _sk, _wh, channel))
                        except Exception: pass
                    conn.commit()
                elif is_outbound:
                    for _r in data_list:
                        try:
                            _sk = _r.get('sku',''); _wh = _r.get('warehouse',''); _qt = int(_r.get('quantity',0) or 0)
                            if _sk and _wh and _qt>0:
                                conn.execute("UPDATE inventory SET month_outbound = COALESCE(month_outbound,0)+? WHERE sku=? AND warehouse=? AND channel=? AND warehouse_type='own'", (_qt, _sk, _wh, channel))
                                # 批次扣减: 出完的行自动删除(不占展开行空间)
                                _pd = str(_r.get('prod_date','') or '')[:10]
                                _ed = str(_r.get('exp_date','') or '')[:10]
                                if _pd and _ed:
                                    conn.execute("UPDATE batches SET qty = MAX(qty - ?, 0) WHERE sku=? AND warehouse=? AND channel=? AND prod_date=? AND exp_date=?", (_qt, _sk, _wh, channel, _pd, _ed))
                                    conn.execute("DELETE FROM batches WHERE sku=? AND warehouse=? AND channel=? AND prod_date=? AND exp_date=? AND qty <= 0", (_sk, _wh, channel, _pd, _ed))
                        except Exception: pass
                    conn.commit()
                # 批次写入: 用同一连接(conn.close()前), 避免后台线程 get_conn thread-local 问题
                if is_inv and batch_to_insert:
                    _keys = set()
                    for _bt in batch_to_insert:
                        _keys.add((_bt[0], _bt[1]))
                    for _k in _keys:
                        try:
                            conn.execute("DELETE FROM batches WHERE sku=? AND warehouse=? AND channel=?", (_k[0], _k[1], channel))
                        except Exception as _be:
                            import logging; logging.warning(f"[cleansing] batch delete: {_be}")
                    for _bt in batch_to_insert:
                        if _bt[3] or _bt[4]:
                            try:
                                conn.execute("INSERT INTO batches(sku, warehouse, warehouse_type, channel, prod_date, exp_date, qty) VALUES(?,?,?,?,?,?,?)", (_bt[0], _bt[1], _bt[2], _bt[3], _bt[4], _bt[5], _bt[6]))
                            except Exception as _be:
                                import logging; logging.warning(f"[cleansing] batch write: {_be}")
                    conn.commit()
                conn.close()
            # 超卖检查：写入的订单量 > 该 SKU 全仓可用库存
            if not is_inv:
                oversell_by_sku = {}
                for o in data_list:
                    sk = o.get('sku','')
                    if sk:
                        oversell_by_sku[sk] = oversell_by_sku.get(sk,0) + int(o.get('quantity',0) or 0)
                # 批量查询库存（避免 N+1：1000 SKU → 1000 次独立查询）
                if oversell_by_sku:
                    skus = list(oversell_by_sku.keys())
                    inv_rows = db.table("inventory").select("*").in_("sku", skus).execute().data or []
                    inv_avail = {}
                    for i in inv_rows:
                        sk = i.get('sku','')
                        inv_avail[sk] = inv_avail.get(sk, 0) + int(i.get('available_qty',0) or 0)
                    for sku, total_qty in oversell_by_sku.items():
                        try:
                            avail = inv_avail.get(sku, 0)
                            if total_qty > avail > 0:
                                db.table("alerts").upsert({
                                    "alert_type":"oversell","title":f"超卖: {sku}","status":"active",
                                    "description":f"导入订单共{total_qty}件 > 可用库存{avail}件",
                                    "severity":"error","source":"cleansing",
                                    "related_sku":sku, "channel": channel,
                                }, conflict_col='alert_type')
                        except Exception as e: print(f"[Cleansing] {e}")
            # 触发事件
            try:
                from app.core.replenishment_cache import invalidate_cache
                invalidate_cache(db)
            except Exception as e: print(f"[Cleansing] {e}")
            # 导入后立即更新日销快照（新订单实时纳入日销计算，不等次日凌晨）
            try:
                from app.core.sales_utils import build_daily_sales_snapshot
                build_daily_sales_snapshot(db)
            except Exception as e: print(f"[Cleansing] {e}")
            # 库存导入后触发规则引擎评估
            if is_inv:
                try:
                    from app.core.rules import evaluate
                    # 批量导入（>100 条）时跳过逐条 evaluate，改为后台批量评估（平衡性能与实时性）
                    if len(data_list) <= 100:
                        for item in data_list:
                            try:
                                evaluate('inventory.changed', {'inv': item, 'db': db, 'sku': item.get('sku',''), 'channel': channel})
                            except Exception as e: print(f"[Cleansing] {e}")
                    else:
                        import logging
                        logging.info(f"[Cleansing] 批量导入 {len(data_list)} 条，提交后台批量规则评估")
                        try:
                            from app.core.database import submit_task
                            from app.api.routes.seed import _seed_rules
                            _tid = f"rule_eval_{datetime.now(timezone.utc).strftime('%H%M%S')}"
                            submit_task(_tid, _seed_rules, get_db(), None, channel=channel, task_type='cleansing')
                        except Exception as e:
                            import logging; logging.warning(f"[Cleansing] batch rule eval: {e}")
                except Exception as e: print(f"[Cleansing] {e}")
            try:
                from app.core.events import bus
                bus.emit('data.cleaned', {
                    'target': target, 'event_type': f'{target}.cleansed',
                    'title': f'清洗导入 {success} 条',
                    'success': success, 'failed': failed,
                    'ws_message': {
                        'type': f'{target}.cleansed',
                        'payload': {'success': success, 'failed': failed}
                    }
                })
            except Exception as e: logger.warning(f"emit event: {e}")
            from app.core.database import submit_task
            from app.api.routes.insights import sync_inventory_from_orders
            submit_task(f"inv_sync_{datetime.now(timezone.utc).strftime('%H%M%S')}", sync_inventory_from_orders, 200)
        except Exception as e:
            from app.core.database import update_task
            update_task(task_id, progress=100)
            return {'ok': False, 'error': f'清洗写入失败: {str(e)[:200]}', 'success': 0, 'failed': 0, 'file': filename, 'target': target}

    msg_parts = []
    if success > 0: msg_parts.append(f"成功导入 {success} 条")
    if failed > 0: msg_parts.append(f"{failed} 条跳过")
    if errors: msg_parts.append(f"{len(errors)} 条异常（可查看错误详情）")

    if success == 0 and failed > 0:
        return {'ok': False, 'success': 0, 'failed': failed, 'file': filename, 'target': target,
                'error': '所有记录均重复或写入失败', 'error_count': len(errors)}
    return {'ok': True, 'success': success, 'failed': failed, 'file': filename,
            'target': target, 'message': '，'.join(msg_parts) if msg_parts else '无数据变更',
            'error_count': len(errors)}

@router.get('/errors')
def get_cleansing_errors(file: str = '', db = get_db()):
    """查询清洗错误记录"""
    if file:
        errs = db.table("cleansing_errors").select("*").eq("source_file", file).order("id", desc=True).limit(500).execute().data
    else:
        errs = db.table("cleansing_errors").select("*").order("id", desc=True).limit(200).execute().data
    for e in errs:
        try: e['raw_data'] = json.loads(e.get('raw_data','{}'))
        except Exception as e: print(f"[Cleansing] {e}")
    return {'ok': True, 'errors': errs, 'total': len(errs)}

@router.post('/execute')
async def execute_cleansing(file: UploadFile = File(...), mapping: str = Form(''),
                             target: str = Form('order'), template_name: str = Form(''), channel: str = 'jd', conflict_mode: str = Form('overwrite')):
    content = await file.read()
    return _run_cleansing(content, file.filename, mapping, target, template_name, channel, conflict_mode)

@router.post('/execute-async')
async def execute_cleansing_async(file: UploadFile = File(...), mapping: str = Form(''),
                                   target: str = Form('order'), template_name: str = Form(''), channel: str = 'jd'):
    import uuid
    content = await file.read()
    # 阈值防护: <400行同步(约<8s PA慢磁盘), >=400行异步(规避30s超时)
    # 行数估算: 数换行符(快, 不完整解析)
    _line_count = content.count(b'\n')
    if _line_count < 400:
        try:
            _res = _run_cleansing(content, file.filename, mapping, target, template_name, channel)
            return _res  # 同步返回结果(success/failed)
        except Exception as _e:
            import logging; logging.error(f"[cleansing] sync execute: {_e}")
            return {'ok': False, 'error': str(_e)[:200], 'success': 0, 'failed': 0}
    task_id = str(uuid.uuid4())[:8]
    submit_task(task_id, _run_cleansing, content, file.filename, mapping, target, template_name, channel, channel=channel, task_type='cleansing')
    from app.core.database import update_task
    update_task(task_id, progress=0, target=target)
    return {'ok': True, 'task_id': task_id, 'message': '任务已提交(异步)'}

@router.get('/templates')
def list_templates_route(db = get_db()):
    return ok(list_templates(db))

@router.post('/templates')
def save_template_route(data: dict, db = get_db()):
    try:
        save_template(data, db)
        return {'ok': True, 'message': '模板已保存'}
    except Exception as e:
        import logging
        logging.error(f"save_template failed: {e}")
        return fail(str(e)[:200])

@router.delete('/templates/{template_id}')
def delete_template_route(template_id: int, db = get_db()):
    delete_template(template_id, db)
    return {'ok': True, 'message': '已删除'}

@router.get('/task/{task_id}')
def get_task_status(task_id: str):
    task = get_task(task_id)
    if not task:
        raise HTTPException(status_code=404, detail='任务不存在')
    return {'ok': True, 'task_id': task_id, **task}

@router.get('/fields/{target}')
def get_fields_route(target: str):
    return ok(get_system_fields(target))

@router.get('/custom-fields/{target}')
def list_custom_fields_route(target: str):
    cf = load_custom_fields()
    return {'ok': True, 'fields': cf.get(target, [])}

@router.post('/custom-fields/{target}')
def add_custom_field_route(target: str, data: dict):
    cf = load_custom_fields()
    if target not in cf: cf[target] = []
    if any(f.get('key') == data.get('key') for f in cf[target]):
        return ok('字段已存在')
    cf[target].append(data)
    save_custom_fields(cf)
    return {'ok': True, 'message': '字段已添加'}

@router.delete('/custom-fields/{target}/{field_key}')
def remove_custom_field_route(target: str, field_key: str):
    cf = load_custom_fields()
    if target in cf:
        cf[target] = [f for f in cf[target] if f.get('key') != field_key]
        save_custom_fields(cf)
    return {'ok': True, 'message': '已删除'}

# ─── 数据库备份 ──────────────────────────────────────────────────────────────

@router.post('/backup')
def trigger_backup():
    path = backup_db()
    if path:
        return {'ok': True, 'path': path, 'message': f'备份完成: {os.path.basename(path)}'}
    return {'ok': False, 'error': '备份失败'}
